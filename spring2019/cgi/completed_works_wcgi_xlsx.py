from openpyxl import load_workbook
import cgi
from wsgiref.simple_server import make_server
import pandas as pd

HTML_PAGE = """<html>
<title>registration</title>
<body>
<h3>Hotel registration</h3>

<form method=POST action=>
<table>
<tr>
<td align=right>

Type of act:
<select name="regim">
<option value="Grounds">Grounds</option>
<option value="Works">Works</option>
<option value="Acts">Acts</option>
<option value="Points">Points</option>
</select>
<br>

Show or add:
<select name="option">
<option value="show">show</option>
<option value="add">add</option>
</select>
<br>


</td>   
<tr>
<td colspan=2 align=center>
<input type=submit value="select">
</td>
<table>
</form>
</body>
</html>
"""


GROUNDS_HTML = """<html lang="en">
<head>
<meta charset="UTF-8">
<title>Completed works</title>
</head>
<body>
<form method=POST action="">
<h1 style="text-align: center">Add grounds</h1>
<p style="text-align: center">
<font size="4" color="blue">fill all entries:</font><br>
<table align="center" style="text-align: center">
<tr>
<td>ID:</td>
<td><input type=text name=s1 value=""></td>
<td>Name:</td>
<td><input type=text name=s2 value=""></td>
<td>Address:</td>
<td><input type=text name=s3 value=""></td>
<td>Responsible:</td>
<td><input type=text name=s4 value=""></td>
<td>Manager:</td>
<td><input type=text name=s5 value=""></td>

</table>
<input type=submit value="Add">
<br>
</form>
</body>
</html>"""



WORKS_HTML = """<html lang="en">
<head>
<meta charset="UTF-8">
<title>completed works</title>
</head>
<body>
<form method=POST action="">
<h1 style="text-align: center">add works</h1>
<p style="text-align: center">
<font size="4" color="blue">fill all entries:</font><br>
<table align="center" style="text-align: center">
<tr>
<td>ID:</td>
<td><input type=text name=s6 value=""></td>
<td>Name:</td>
<td><input type=text name=s7 value=""></td>

</table>
<input type=submit value="Add">
<br>
</form>
</body>
</html>"""

ACTS_HTML = """<html lang="en">
<head>
<meta charset="UTF-8">
<title>Completed works</title>
</head>
<body>
<form method=POST action="">
<h1 style="text-align: center">Add acts</h1>
<p style="text-align: center">
<font size="4" color="blue">fill all entries:</font><br>
<table align="center" style="text-align: center">
<tr>
<td>ID:</td>
<td><input type=text name=s8 value=""></td>
<td>No:</td>
<td><input type=text name=s9 value=""></td>
<td>Date:</td>
<td><input type=text name=s10 value=""></td>
<td>Sum:</td>
<td><input type=text name=s11 value=""></td>
<td>S_id:</td>
<td><input type=text name=s12 value=""></td>

</table>
<input type=submit value="Add">
<br>
</form>
</body>
</html>"""

POINTS_HTML = """<html lang="en">
<head>
<meta charset="UTF-8">
<title>completed works</title>
</head>
<body>
<form method=POST action="">
<h1 style="text-align: center">all ponts</h1>
<p style="text-align: center">
<font size="4" color="blue">fill all entries:</font><br>
<table align="center" style="text-align: center">
<tr>
<td>W_id:</td>
<td><input type=text name=s13 value=""></td>
<td>A_id:</td>
<td><input type=text name=s14 value=""></td>
</table>
<input type=submit value="Add">
<br>
</form>
</body>
</html>"""

def show_xml(sheet, filename="info.xlsx"):
    html_df = ""
    df = pd.read_excel(filename, sheet_name=sheet)
    for row in df.values:
        html_df += "\n<br>"
        for val in row:
            html_df += str(val) +" "
    return html_df

def to_xml(array, sheet, filename="info.xlsx"):
    wb = load_workbook(filename)
    ws = ""
    try:
        ws = wb[sheet]
    except:
        wb.create_sheet(sheet)
        ws = wb[sheet]

    ws.append(array)
    wb.save(filename=filename)


state = 0


def application(environ, start_response):

    global state

    if environ.get('PATH_INFO', '').lstrip('/') == '':
        form = cgi.FieldStorage(fp=environ['wsgi.input'], environ=environ)

        html = HTML_PAGE

        if state == 1:
            if 's1' in form and 's2' in form and 's3' in form and 's4' in form and 's5' in form:
                to_xml([str(form['s1'].value), str(form['s2'].value), str(form['s3'].value),
                        str(form['s4'].value), str(form['s5'].value)], "grounds")

        elif state == 2:
            if 's6' in form and 's7' in form:
                to_xml([str(form['s6'].value), str(form['s7'].value)], "works")


        elif state == 3:
            if 's8' in form and 's9' in form and 's10' in form and 's11' in form and 's12' in form:
                to_xml([str(form['s8'].value), str(form['s9'].value), str(form['s10'].value),
                        str(form['s11'].value), str(form['s12'].value)], "acts")

        elif state == 4:
            if 's13' in form and 's14' in form:
                to_xml([str(form['s13'].value), str(form['s14'].value)], "points")
                
        else:
            html = HTML_PAGE

        if 'regim' in form:
            val = form['regim'].value
            option = form['option'].value
            if option == "add":
                if val == 'Grounds':
                    state = 1
                    html = GROUNDS_HTML

                elif val == 'Works':
                    state = 2
                    html = WORKS_HTML

                elif val == 'Acts':
                    state = 3
                    html = ACTS_HTML

                elif val == 'Points':
                    state = 4
                    html = POINTS_HTML

            else:

                if val == 'Grounds':
                    state = 5
                    html = show_xml("grounds")

                elif val == 'Works':
                    state = 6
                    html = show_xml("works")

                elif val == 'Acts':
                    state = 7
                    html = show_xml("acts")

                elif val == 'Points':
                    state = 8
                    html = show_xml("points")

        else:
            state = 0


        body = html
        start_response('200 OK', [('Content-Type', 'text/html; charset=utf-8')])

    else:

        start_response('404 NOT FOUND', [('Content-Type', 'text/plain; charset=utf-8')])
        body = 'Сторінку не знайдено'

    return [bytes(body, encoding='utf-8')]


print('=== Local WSGI webserver ===')
httpd = make_server('localhost', 8090, application)
httpd.serve_forever()

