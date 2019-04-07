import urllib.request
import json
import openpyxl
from openpyxl import load_workbook


class CurencyConverter:

    def __init__(self, path_to_download):
        self.cur_array = []
        self.rez = 0
        self.list_of_currency()
        self.final_write()
        self.path_to_download = path_to_download

    #
    def rez_lst(self, key1, key2):
        url_prt1 = "https://free.currencyconverterapi.com/api/v6/convert?q="
        url_prt2 = "&compact=ultra&apiKey=782dddca2acbac7d3d84"
        key = key1 + "_" + key2
        print(key)
        url = url_prt1 + key + url_prt2
        req = urllib.request.Request(url)

        data = urllib.request.urlopen(req).read()

        data = json.loads(data.decode("utf-8"))

        return data[key]


    def list_of_currency(self):
        main_cur = ["USD", "UAH"] #, "EUR", "RUB", "PLN", "CHF", "GBP"]
        lenth = len(main_cur)
        for i in range(lenth):
            for j in range(lenth):
                self.cur_array.append([main_cur[i], main_cur[j]])


    def final_write(self):
        for i in range(len(self.cur_array)):
            self.cur_array[i].append(self.rez_lst(self.cur_array[i][0], self.cur_array[i][1]))
        print(self.cur_array)

        self.write_xlsx(self.cur_array)



    def write_xlsx(self, array):
        wb = load_workbook("/home/stas/Documents/currency.xlsx")
        ws = wb.create_sheet(0)
        for row in array:
            ws.append(row)

        wb.save(filename="/home/stas/Documents/currency.xlsx")


CurencyConverter(path_to_download="/home/stas/Documents/currency.xlsx")